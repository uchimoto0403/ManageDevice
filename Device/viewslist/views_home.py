from django.shortcuts import render, redirect
from django.urls import reverse
from Device.models import UserMst, DeviceMst
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import logging
#------------------------------------------------------------------------------------------------#

#ホーム_顧客
#引　数：リクエスト　ユーザー種類
#戻り値：なし

def home_customer(request, struserid):
    try:
         
        #不正アクセスが起きた場合
        objuser = UserMst.objects.filter(id = struserid)     
        if objuser.count() <= 0 : 
            
            # ログイン画面に移行
            request.session.flush()
            strurl = reverse( 'login' )
            return redirect( strurl )
        
        objuser = UserMst.objects.get(id = struserid)
        devices = DeviceMst.objects.filter( dvcCustomer = objuser )  

        # 共通パラメータ定義
        params = {
            'User'          : objuser,
            'devices'       : devices,
            'struserid'     : struserid,
        }
         
        # GET時処理
        if request.method == 'GET':

            # ホーム画面表示
            return render( request, 'Home_Customer.html', params )    
        
        # POST時処理
        if request.method == 'POST':

            # 詳細ボタン押下時
            if 'btnDetail' in request.POST:
                # 詳細画面に移行
                strurl = reverse('detail_device', kwargs={'struserid': struserid, 'strdevid': request.POST['btnDetail']})
                return redirect(strurl)
            
            # 出力ボタン押下時
            elif 'btnOutput' in request.POST:

                # Excelファイル作成
                wb = load_workbook(r"C:\Users\PC1-30_uchimoto\Desktop\Python 危機管理システム\05.コーディング\機器一覧出力_顧客用.xlsx")
                ws = wb["Sheet1"]

                ws["G3"] = objuser.usrCustomer
                
                # 登録されている機器情報取得
                devices = DeviceMst.objects.filter(dvcCustomer=objuser)

                # 開始位置を D7 に設定
                start_row = 7
                start_col = 4  

                for idx, device in enumerate(devices):
                    row_num = start_row + idx

                    softwares = device.dvs_dvc_id.all()
                    sw_names = ", ".join([sw.dvsSoftName for sw in softwares]) if softwares else ""
                    sw_warranties = ", ".join([sw.dvsWarranty.strftime("%Y-%m-%d") for sw in softwares]) if softwares else ""

                    # 機器情報とソフト情報を1行にまとめる
                    values = [
                        device.dvcName,
                        device.dvcKind,
                        device.dvcMaker,
                        device.dvcPurchase,
                        device.dvcWarranty,
                        device.dvcUser,
                        device.dvcPlace,
                        sw_names,
                        sw_warranties,  
                        device.dvcNotes
                    ]

                    for col_offset, value in enumerate(values):
                        ws.cell(row=row_num, column=start_col + col_offset, value=value)

                # HttpResponseを使用してExcelファイルを返す
                from django.http import HttpResponse       
                response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename=Device_List_{objuser.usrCustomer}.xlsx'
                wb.save(response)
                return response
            
            # ログアウトボタン押下時
            elif 'btnLogout' in request.POST:

                # ログイン画面に移行
                strurl = reverse( 'login' )
                return redirect( strurl )
            
        return render(request, 'home_customer.html', params)


    except:
        # トレース設定
        import traceback

        # ログ出力
        logger = logging.getLogger(__name__)
        logger.error( request )
        logger.error( traceback.format_exc() )

    return render(request, 'home_customer.html', params) 


#ホーム_管理者
#引　数：リクエスト　ユーザー種類
#戻り値：なし

def home_admin( request, struserid ):
    
    try:
       
        #不正アクセスが起きた場合
        objuser = UserMst.objects.filter( id = struserid )         
        if objuser.count() <= 0 : 

            # ログイン画面に移行
            request.session.flush()
            return redirect( 'login' )
        
        objuser = UserMst.objects.get( id = struserid )
        
        # 共通パラメータ定義
        params = {
            'User'          : objuser,          
            }

        # GET時処理
        if request.method == 'GET':

            # ホーム画面表示
            return render( request, 'home_admin.html', params )

        # POST時処理
        if request.method == 'POST':

            # 顧客管理ボタン押下時
            if 'btnCustomer' in request.POST:

                # 顧客管理画面に移行
                strurl = reverse( 'manage_customer', kwargs = { 'struserid' : objuser.id } )
                return redirect( strurl )  
            
            # 機器管理ボタン押下時
            elif 'btnDevice' in request.POST:

                # 機器管理画面に移行
                strurl = reverse( 'manage_device', kwargs = { 'struserid' : objuser.id } )
                return redirect( strurl )  
            
            # 管理者管理ボタン押下時
            elif 'btnAdmin' in request.POST:

                # 管理者管理画面に移行
                strurl = reverse( 'manage_admin', kwargs = { 'struserid' : objuser.id } )
                return redirect( strurl )  
            
            # ログアウトボタン押下時
            elif 'btnLogout' in request.POST:

                # ログイン画面に移行
                strurl = reverse( 'login' )
                return redirect( strurl )

        return render(request, 'Home_Admin.html', params)     
         
    except:
        # トレース設定
        import traceback

        # ログ出力
        logger = logging.getLogger(__name__)
        logger.error( request )
        logger.error( traceback.format_exc() )

    return render(request, 'Home_Admin.html', params) 


    

        
