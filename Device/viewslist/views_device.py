from django.shortcuts import render, redirect
from django.urls import reverse
from django.db import transaction
from Device.models import UserMst ,DeviceMst, DeviceSoftMst
from Device.formslist.forms_device import DeviceForm
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from django.contrib import messages
import logging
import io
from django.contrib.staticfiles import finders
#------------------------------------------------------------------------------------------------#

  

# 機器管理
# 引　数：リクエスト　ユーザーID 機器ID
# 戻り値：なし

def manage_device(request, struserid):
    try:

        #不正アクセスが起きた場合
        objuser = UserMst.objects.filter(id=struserid)
        if objuser.count() <= 0:

            # ログイン画面に移行
            request.session.flush()
            strurl = reverse( 'login' )
            return redirect( strurl )
        
        
        objuser     = UserMst.objects.get(id=struserid)
        customers   = UserMst.objects.filter(usrKind=1, usrDelete=False)
        blnerror    = False
        blnerror_c  = False


        # 共通パラメータ定義
        params = {
            'User'                  : objuser,
            'struserid'             : struserid,            # ユーザーID
            'customers'             : customers,            # 顧客情報
            'Required_error'        : blnerror,
           }

        # GET時処理
        if request.method == 'GET':

            customers = UserMst.objects.filter(usrKind=1, usrDelete=False)

            # 機器管理画面表示
            return render( request, 'manage_device.html', params ) 
        
        # POST時処理
        if request.method == 'POST':

            # 機器登録ボタン押下時
            if 'btnCreate' in request.POST:

                # 機器登録画面に移行
                strurl = reverse( 'create_device', kwargs = { 'struserid' : objuser.id } )
                return redirect( strurl ) 
             
            # 検索ボタン押下時
            if 'btnSearch' in request.POST:
                customer_id = request.POST.get('customer_id')

                # 顧客を選択していない場合
                if not customer_id:

                    blnerror = True  

                    # パラメータ更新
                    params['RequiredError'] = blnerror

                    return render( request, 'manage_device.html', params )
             
            # 顧客情報取得
                customer = UserMst.objects.filter(id=customer_id, usrKind=1, usrDelete=False).first()
                if not customer:
                    blnerror = True
                    params['RequiredError'] = blnerror
                    return render(request, 'manage_device.html', params)

                # 機器情報取得
                devices = DeviceMst.objects.filter(dvcCustomer=customer, dvcDeleteFlag=False)
                for device in devices:
                    device.has_software = DeviceSoftMst.objects.filter(dvsDeviceID=device, dvsDeleteFlag=False).exists()                

                if not devices.exists():
                    blnerror = True
                    params['RequiredError'] = blnerror
                    return render(request, 'manage_device.html', params)

                # 顧客と機器リストを返す
                params['Devices'] = devices
                params['SelectedCustomer'] = customer
                params['can_output'] = devices.exists()
                return render(request, 'manage_device.html', params)

            # ソフト確認ボタン押下時
            if 'btnCheckSoftware' in request.POST:
                device_id = request.POST.get('btnCheckSoftware')
                device = DeviceMst.objects.get(id=device_id)

                # ソフト一覧を取得
                softwares = DeviceSoftMst.objects.filter(dvsDeviceID=device, dvsDeleteFlag=False)

                customer = device.dvcCustomer

                devices = DeviceMst.objects.filter(dvcCustomer=customer, dvcDeleteFlag=False)
                for device in devices:
                    device.has_software = DeviceSoftMst.objects.filter(dvsDeviceID=device, dvsDeleteFlag=False).exists()

                params['Devices'] = devices
                params['softwares'] = softwares
                params['selected_device'] = device
                params['SelectedCustomer'] = customer 
                params['open_modal'] = True
                params['can_output'] = devices.exists()

                if not softwares.exists():
                    params['error_software'] = True
                else:
                    params['softwares'] = softwares

                return render(request, 'manage_device.html', params) 

            # 編集ボタン押下時
            if 'btnEdit' in request.POST:
                device_id = request.POST.get('btnEdit')

                # URLに必要なのは strdevid
                strurl = reverse('edit_device', kwargs={'struserid': objuser.id, 'strdevid': device_id})
                return redirect(strurl)


            # 削除ボタン押下時
            if 'btnDelete' in request.POST:
                device_id = request.POST.get('btnDelete')

                with transaction.atomic():
                    # 削除対象の機器を取得
                    device = DeviceMst.objects.get(id=device_id)

                    # 削除フラグを立てて保存
                    device.dvcDeleteFlag = True
                    device.save()

                # 再検索してリスト更新（削除後にリストから消すため）
                customer = device.dvcCustomer
                devices = DeviceMst.objects.filter(dvcCustomer=customer, dvcDeleteFlag=False)
                for d in devices:
                    d.has_software = DeviceSoftMst.objects.filter(dvsDeviceID=d, dvsDeleteFlag=False).exists()

                params['Devices'] = devices
                params['SelectedCustomer'] = customer
                params['can_output'] = devices.exists()

                return render(request, 'manage_device.html', params)
            
            # 出力ボタン押下時
            if 'btnOutput' in request.POST:
                # どの顧客を出力するか取得（検索で選んだ顧客）
                selected_customer_id = request.POST.get('selected_customer_id')

                if not selected_customer_id:
                    messages.error(request, "出力対象の顧客が選択されていません。先に顧客を検索してください。")
                    return render(request, 'manage_device.html', params)

                customer = UserMst.objects.filter(id=selected_customer_id, usrKind=1, usrDelete=False).first()
                if not customer:
                    messages.error(request, "出力対象の顧客が存在しません。")
                    return render(request, 'manage_device.html', params)

                # static 内の Excel テンプレートを探す
                template_path = finders.find("excel/機器一覧出力_管理者用.xlsx")
                if not template_path:
                   messages.error(request, "Excelテンプレートが見つかりません")
                   return render(request, 'manage_device.html', params)
                # Excelファイル作成
                wb = load_workbook(template_path)
                ws = wb["Sheet1"]

                ws["G3"] = objuser.usrCustomer
                
                # 登録されている機器情報取得
                devices = DeviceMst.objects.filter(dvcCustomer=customer, dvcDeleteFlag=False).order_by('id')

                    # もしデータが無ければメッセージを返す
                if not devices.exists():
                    messages.error(request, "該当顧客の機器データがありません。")
                    return render(request, 'manage_device.html', params)

            # 開始位置を D7 に設定
                start_row = 7
                start_col = 4  

                for idx, device in enumerate(devices):
                    row_num = start_row + idx

                    softwares = device.dvs_dvc_id.filter(dvsDeleteFlag=False)
                    sw_names = ", ".join([sw.dvsSoftName for sw in softwares]) if softwares else ""
                    sw_warranties = ", ".join([sw.dvsWarranty.strftime("%Y-%m-%d") for sw in softwares]) if softwares else ""

                    # 機器情報とソフト情報を1行にまとめる
                    values = [
                        device.dvcName,
                        device.dvcKind,
                        device.dvcMaker,
                        device.dvcPurchase,
                        device.dvcWarranty,
                        device.dvcUser,
                        device.dvcPlace,
                        device.dvcAssetnumber,
                        device.dvcStatus,
                        device.dvcSerialnumber,
                        device.dvcOS,
                        device.dvcCPU,
                        device.dvcRAM,
                        device.dvcGraphic,
                        device.dvcStorage,
                        device.dvcIP,
                        device.dvcNetWork,
                        sw_names,
                        sw_warranties,
                        device.dvcNotes,  
                    ]

                    for col_offset, value in enumerate(values):
                        ws.cell(row=row_num, column=start_col + col_offset, value=value)

                # ===== 自動列幅調整 =====
                for col in ws.columns:
                    max_length = 0
                    col_letter = get_column_letter(col[0].column)
                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 2  # 余白を持たせる
                    ws.column_dimensions[col_letter].width = adjusted_width


                import io
                from django.http import HttpResponse

                output = io.BytesIO()
                wb.save(output)
                output.seek(0)

                response = HttpResponse(
                    output,
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                response['Content-Disposition'] = f'attachment; filename=Device_List_{customer.usrCustomer}.xlsx'
                return response
  

          
            # 戻るボタン押下時
            elif 'btnBack' in request.POST:

                # ホーム_管理者画面に移行
                strurl = reverse( 'home_admin', kwargs = { 'struserid' : struserid } )
                return redirect( strurl )
            
            # ログアウトボタン押下時
            elif 'btnLogout' in request.POST:

                # ログイン画面に移行
                return redirect( 'login' )

    except:
        # トレース設定
        import traceback

        # ログ出力
        logger = logging.getLogger(__name__)
        logger.error( request )
        logger.error( traceback.format_exc() )

        # ログイン画面に移行
        strurl = reverse( 'login' )
        return redirect( strurl )
    
    return render(request, 'manage_Device.html', params)            

# 詳細
# 引　数：リクエスト　ユーザーID　機器ID
# 戻り値：なし

def detail_device(request, struserid, strdevid ):
    try:
         
        #不正アクセスが起きた場合
        objuser = UserMst.objects.filter(id=struserid)      
        if objuser.count() <= 0 : 
            
            # ログイン画面に移行
            request.session.flush()
            strurl = reverse( 'login' )
            return redirect( strurl )
        
        params = {} 
        
        # 引数で渡すものを指定
        objuser = UserMst.objects.get(id=struserid)
        device  = DeviceMst.objects.get( id = strdevid )
        devicesofts = DeviceSoftMst.objects.filter(dvsDeviceID = device, dvsDeleteFlag=False )

        # 共通パラメータ定義
        params = {
            'User'                      : objuser,
            'device'                    : device,  
            'devicesofts'               : devicesofts, 
            'struserid'                 : struserid,           # ユーザーID
            "strdevid"                  : strdevid,            # 機器ID
            }
         
        # GET時処理
        if request.method == 'GET':

            # ホーム画面表示
            return render( request, 'detail_device.html', params )    
        
        # POST時処理
        if request.method == 'POST':

            # ソフト詳細ボタン押下時
            if 'btnCheck' in request.POST:
                return render( request, 'detail_device.html', params )

            # 戻るボタン押下時
            if 'btnBack' in request.POST:
                # ホーム画面に移行
                strurl = reverse('home_customer', kwargs={'struserid': struserid})
                return redirect(strurl)
            
            # ログアウトボタン押下時
            elif 'btnLogout' in request.POST:

                # ログイン画面に移行
                strurl = reverse( 'login' )
                return redirect( strurl )
            
        return render(request, 'detail_device.html', params)
        
    except:
        # トレース設定
        import traceback

        # ログ出力
        logger = logging.getLogger(__name__)
        logger.error( request )
        logger.error( traceback.format_exc() )

    return render( request, 'detail_device.html', params )


# 機器登録
# 引　数：リクエスト　ユーザーID
# 戻り値：なし

def create_device(request, struserid):
    try:
        objuser = UserMst.objects.filter(id=struserid)
        if not objuser.exists():
            request.session.flush()
            return redirect('login')

        objuser = objuser.first()
        customers = UserMst.objects.filter(usrKind=1, usrDelete=False)

        temp_softs = request.session.get('temp_softs', [])
        temp_device = request.session.get('temp_device', {})

        params = {
            'User': objuser,
            'Form': DeviceForm(),
            'struserid': struserid,
            'customers': customers,
            'temp_softs': temp_softs,
            'temp_device': temp_device,
        }

        if request.method == 'GET':
            return render(request, 'create_device.html', params)

        if request.method == 'POST':

            # 機器登録
            if 'btnCreate' in request.POST:
                device_name = request.POST.get('chrDeviceName', '').strip()
                customer_id = request.POST.get('intCustomer')

                if not device_name or not customer_id:
                    messages.error(request, "機器名と顧客は必須です")
                    request.session['temp_device'] = request.POST.dict()
                    params['temp_device'] = request.POST.dict()
                    return render(request, 'create_device.html', params)

                customer = UserMst.objects.filter(id=customer_id, usrKind=1, usrDelete=False).first()
                if not customer:
                    messages.error(request, "顧客が存在しません")
                    request.session['temp_device'] = request.POST.dict()
                    params['temp_device'] = request.POST.dict()
                    return render(request, 'create_device.html', params)

                device = DeviceMst.objects.create(
                    dvcName=device_name,
                    dvcCustomer=customer,
                    dvcKind=request.POST.get('chrDeviceKind', ''),
                    dvcMaker=request.POST.get('chrDeviceMaker', ''),
                    dvcModel=request.POST.get('chrDeviceModel', ''),
                    dvcPurchase=request.POST.get('dtDevicePurchase') or None,
                    dvcWarranty=request.POST.get('dtDeviceWarranty') or None,
                    dvcUser=request.POST.get('chrDeviceUser', ''),
                    dvcPlace=request.POST.get('chrDevicePlace', ''),
                    dvcAssetnumber=request.POST.get('chrDeviceAssetNumber', ''),
                    dvcStatus=request.POST.get('chrDeviceStatus', ''),
                    dvcSerialnumber=request.POST.get('chrDeviceSerialNumber', ''),
                    dvcOS=request.POST.get('chrDeviceOS', ''),
                    dvcCPU=request.POST.get('chrDeviceCPU', ''),
                    dvcRAM=request.POST.get('chrDeviceRAM', ''),
                    dvcGraphic=request.POST.get('chrDeviceGraphic', ''),
                    dvcStorage=request.POST.get('chrDeviceStorage', ''),
                    dvcIP=request.POST.get('chrDeviceIP', ''),
                    dvcNetWork=request.POST.get('chrDeviceNetwork', ''),
                    dvcNotes=request.POST.get('chrNotes', ''),
                    dvcDeleteFlag=False,
                )

                for soft in temp_softs:
                    DeviceSoftMst.objects.create(
                        dvsDeviceID=device,
                        dvsSoftName=soft['name'],
                        dvsWarranty=soft['warranty'],
                        dvsDeleteFlag=False
                    )

                request.session['temp_softs'] = []
                request.session.pop('temp_device', None)

                messages.success(request, "機器とソフトを登録しました")
                return redirect('create_device', struserid=struserid)

            # ソフト追加
            if 'btnAddSoftTemp' in request.POST:
                request.session['temp_device'] = request.POST.dict()
                name = request.POST.get('chrSoftName', '').strip()
                warranty = request.POST.get('chrWarranty', '').strip()
                if name and warranty:
                    temp_softs.append({'name': name, 'warranty': warranty})
                    request.session['temp_softs'] = temp_softs
                    messages.success(request, f"ソフト '{name}' を追加しました")
                else:
                    messages.error(request, "ソフト名と保証期限を入力してください")
                params['temp_softs'] = temp_softs
                params['temp_device'] = request.session['temp_device']
                return render(request, 'create_device.html', params)

            # ソフト編集・削除
            if 'btnUpdateSoftTemp' in request.POST or 'btnDeleteSoftTemp' in request.POST:
                request.session['temp_device'] = request.POST.dict()

                if 'btnUpdateSoftTemp' in request.POST:
                    index = int(request.POST.get('soft_index'))
                    new_name = request.POST.get('chrSoftName', '').strip()
                    new_warranty = request.POST.get('chrWarranty', '').strip()
                    if 0 <= index < len(temp_softs):
                        if new_name and new_warranty:
                            temp_softs[index]['name'] = new_name
                            temp_softs[index]['warranty'] = new_warranty
                            request.session['temp_softs'] = temp_softs
                            messages.success(request, "ソフトを更新しました")
                        else:
                            messages.error(request, "ソフト名と保証期限を入力してください")

                if 'btnDeleteSoftTemp' in request.POST:
                    index = int(request.POST.get('soft_index'))
                    if 0 <= index < len(temp_softs):
                        deleted = temp_softs.pop(index)
                        request.session['temp_softs'] = temp_softs
                        messages.success(request, f"ソフト '{deleted['name']}' を削除しました")

                params['temp_softs'] = temp_softs
                params['temp_device'] = request.session['temp_device']
                return render(request, 'create_device.html', params)

            # 戻る・ログアウト
            if 'btnBack' in request.POST:
                request.session['temp_softs'] = []
                request.session.pop('temp_device', None)
                return redirect(reverse('manage_device', kwargs={'struserid': struserid}))

            if 'btnLogout' in request.POST:
                request.session['temp_softs'] = []
                request.session.pop('temp_device', None)
                return redirect('login')

        return render(request, 'create_device.html', params)

    except Exception:
        import traceback
        logger = logging.getLogger(__name__)
        logger.error(traceback.format_exc())
        return redirect('login')

# 機器編集
# 引　数：リクエスト　ユーザーID　機器ID
# 戻り値：なし

# 機器編集画面（edit_device）
def edit_device(request, struserid, strdevid):
    try:
        # --- ログインユーザー確認 ---
        objuser = UserMst.objects.filter(id=struserid).first()
        if not objuser:
            request.session.flush()
            return redirect('login')

        # --- 顧客一覧と対象機器 ---
        customers = UserMst.objects.filter(usrKind=1, usrDelete=False)
        device = DeviceMst.objects.get(id=strdevid, dvcDeleteFlag=False)

        # --- 一時セッション内ソフト情報 ---
        temp_softs = request.session.get('temp_softs', [])
        softwares = DeviceSoftMst.objects.filter(dvsDeviceID=device, dvsDeleteFlag=False)

        # 初回表示時、一時データが空なら既存ソフトを反映
        if not temp_softs and softwares.exists():
            temp_softs = [
                {'name': s.dvsSoftName, 'warranty': s.dvsWarranty.strftime('%Y-%m-%d') if s.dvsWarranty else ''}
                for s in softwares
            ]
            request.session['temp_softs'] = temp_softs

        params = {
            'User': objuser,
            'struserid': struserid,
            'device': device,
            'customers': customers,
            'temp_softs': temp_softs,
            'open_modal': False,  # 初回ではモーダル開かない
        }
      
        # GET（初回表示）       
        if request.method == 'GET':
            form = DeviceForm(initial={
                'chrDeviceName': device.dvcName,
                'chrDeviceKind': device.dvcKind,
                'chrDeviceMaker': device.dvcMaker,
                'chrDeviceModel': device.dvcModel,
                'dtDevicePurchase': device.dvcPurchase,
                'dtDeviceWarranty': device.dvcWarranty,
                'chrDeviceUser': device.dvcUser,
                'chrDevicePlace': device.dvcPlace,
                'chrDeviceAssetNumber': device.dvcAssetnumber,
                'chrDeviceStatus': device.dvcStatus,
                'chrDeviceSerialNumber': device.dvcSerialnumber,
                'chrDeviceOS': device.dvcOS,
                'chrDeviceCPU': device.dvcCPU,
                'chrDeviceRAM': device.dvcRAM,
                'chrDeviceGraphic': device.dvcGraphic,
                'chrDeviceStorage': device.dvcStorage,
                'chrDeviceIP': device.dvcIP,
                'chrDeviceNetwork': device.dvcNetWork,
                'chrNotes': device.dvcNotes,
            })
            params['Form'] = form
            return render(request, 'edit_device.html', params)

        
        # POST共通        
        form = DeviceForm(request.POST)
        params['Form'] = form

        # ===== ソフト追加 =====
        if 'btnAddSoftTemp' in request.POST:
            name = request.POST.get('chrSoftName', '').strip()
            warranty = request.POST.get('chrWarranty', '').strip()
            if name and warranty:
                temp_softs.append({'name': name, 'warranty': warranty})
                request.session['temp_softs'] = temp_softs
                messages.success(request, f"ソフト '{name}' を追加しました。")
            else:
                messages.error(request, "ソフト名と保証期限を入力してください。")

            params['temp_softs'] = temp_softs
            params['open_modal'] = True
            return render(request, 'edit_device.html', params)

        # ===== ソフト更新 =====
        if 'btnUpdateSoftTemp' in request.POST:
            index = int(request.POST.get('soft_index', -1))
            new_name = request.POST.get('chrSoftName', '').strip()
            new_warranty = request.POST.get('chrWarranty', '').strip()
            if 0 <= index < len(temp_softs):
                temp_softs[index]['name'] = new_name
                temp_softs[index]['warranty'] = new_warranty
                request.session['temp_softs'] = temp_softs
                messages.success(request, "ソフト情報を更新しました。")
            params['open_modal'] = True
            return render(request, 'edit_device.html', params)

        # ===== ソフト削除 =====
        if 'btnDeleteSoftTemp' in request.POST:
            index = int(request.POST.get('soft_index', -1))
            if 0 <= index < len(temp_softs):
                deleted = temp_softs.pop(index)
                request.session['temp_softs'] = temp_softs
                messages.success(request, f"ソフト '{deleted['name']}' を削除しました。")
            params['open_modal'] = True
            return render(request, 'edit_device.html', params)

        # ===== 機器更新 =====
        if 'btnUpdateDevice' in request.POST:
            if form.is_valid():
                cd = form.cleaned_data
                device.dvcName = cd['chrDeviceName']
                device.dvcKind = cd['chrDeviceKind']
                device.dvcMaker = cd['chrDeviceMaker']
                device.dvcModel = cd['chrDeviceModel']
                device.dvcPurchase = cd['dtDevicePurchase']
                device.dvcWarranty = cd['dtDeviceWarranty']
                device.dvcUser = cd['chrDeviceUser']
                device.dvcPlace = cd['chrDevicePlace']
                device.dvcAssetnumber = cd['chrDeviceAssetNumber']
                device.dvcStatus = cd['chrDeviceStatus']
                device.dvcSerialnumber = cd['chrDeviceSerialNumber']
                device.dvcOS = cd['chrDeviceOS']
                device.dvcCPU = cd['chrDeviceCPU']
                device.dvcRAM = cd['chrDeviceRAM']
                device.dvcGraphic = cd['chrDeviceGraphic']
                device.dvcStorage = cd['chrDeviceStorage']
                device.dvcIP = cd['chrDeviceIP']
                device.dvcNetWork = cd['chrDeviceNetwork']
                device.dvcNotes = cd['chrNotes']
                device.save()

                # ソフト更新（既存削除→新規登録）
                DeviceSoftMst.objects.filter(dvsDeviceID=device).update(dvsDeleteFlag=True)
                for soft in temp_softs:
                    DeviceSoftMst.objects.create(
                        dvsDeviceID=device,
                        dvsSoftName=soft['name'],
                        dvsWarranty=soft['warranty'] or None,
                        dvsDeleteFlag=False
                    )

                request.session['temp_softs'] = []
                messages.success(request, "機器情報を更新しました。")
                return redirect('manage_device', struserid=struserid)
            else:
                messages.error(request, "入力内容に誤りがあります。")
                return render(request, 'edit_device.html', params)

        # ===== 戻る =====
        if 'btnBack' in request.POST:
            request.session['temp_softs'] = []
            return redirect('manage_device', struserid=struserid)

        # ===== ログアウト =====
        if 'btnLogout' in request.POST:
            request.session.flush()
            return redirect('login')

        return render(request, 'edit_device.html', params)

    except Exception:
        import traceback
        logging.getLogger(__name__).error(traceback.format_exc())
        return redirect('login')
